from django.shortcuts import render,get_object_or_404
from django.views.generic import ListView
from django.http import HttpResponse,HttpResponseRedirect,JsonResponse
from django.urls import reverse
from dashboard.models import *
from django.core.files.storage import FileSystemStorage
from openpyxl import Workbook,load_workbook
from django.contrib.auth.models import User
from django.contrib.auth import logout,login,authenticate
from collections import defaultdict
import copy
import json
# Create your views here.
def dashboard(request):
    i = Invoice.objects.all()
    print(len(i))
    for j in i:
        i.delete()
    print(Invoice.objects.all())
    return render(request,'index.html',{})

def itemView(request):
    pk = 0
    item = Item.objects.get(pk)
    return render(request, '',{'item':item})

def invoiceView(request):
    pk = 0
    invoice = Invoice.objects.all()
    return render(request, 'invoice.html', {'invoices':invoice})

def inven(request):
    inventory = Item.objects.all()
    return render(request, 'inv.html', {'items':inventory})

def invoiceDetails(request, pk):
    invoice = Invoice.objects.get(pk=pk)

    return render(request, 'invoicedetails.html', {'invoice': invoice})

def addInvoice(request):
    items = Item.objects.all()
    return render(request,"invoiceAdd.html",{"items":items})

def getQuantity(request):
    pk = request.POST["pk"]
    item = Item.objects.get(pk=pk)
    return JsonResponse({"quantity":item.stock})

def createInvoice(request):
    if request.method == "POST":
        invoice = Invoice()
        invoice.customerNum = "99"
        invoice.customerfName = request.POST["customerFirstName"]
        invoice.customerlName = request.POST["customerLastName"]
        invoice.total = request.POST["total"]
        invoice.discount = request.POST["discount"]
        try:
            invoice.sendInvoice = request.POST["sendInvoice"]
        except:
            invoice.sendInvoice = False
        invoice.finalSale = request.POST["finalSale"]
        invoice.save()

        items = []
        keys = request.POST.keys()
        seenKeys = []
        for key in request.POST.keys():
            if ("invoicePK" in key or "invoiceQuantity" in key) and not key in seenKeys:
                if "invoicePK" in key:
                    items.append(tuple([request.POST[key],request.POST["invoiceQuantity" + key[-1]]]))
                else:
                    items.append(tuple([request.POST["invoicePK" + key[-1]],request.POST[key]]))
                seenKeys.append(key)
                seenKeys.append("invoiceQuantity" + key[-1])

        for item in items:
            invoiceItem = InvoiceItem()
            invoiceItem.item = Item.objects.get(pk=item[0])
            invoiceItem.quantity = item[1]
            invoiceItem.invoice = invoice
            invoiceItem.save()

        return HttpResponseRedirect(reverse("invoiceView"))

def addInventoryManual(request,addAnother=None):
    if request.method=="GET":
        items = Item.objects.all()
        return render(request,"addManualStock.html",{"items":items})
    else:
        pk = request.POST["item"]
        item = Item.objects.get(pk=pk)
        item.stock += int(request.POST["count"])
        item.save()

        update = stockUpdate()
        update.item = item
        update.count = request.POST["count"]
        update.save()

        if addAnother:
            return HttpResponseRedirect(reverse("addIneventoryManual"))
        else:
            return HttpResponseRedirect(reverse("inventoryView"))

def addExcelInventory(request):

    if request.method=="POST" and request.FILES:

        file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(file.name,file)
        url = fs.url(filename)

        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_col=0, max_col=2):
            if row[0].value == "Item (Do not touch number between [])":
                continue
            item_pk = row[0].value.split("[")[1][:-1]
            count = row[1].value

            item = Item.objects.get(pk=item_pk)
            item.stock += count
            item.save()

            update = stockUpdate()
            update.item = item
            update.count = count
            update.save()


        return HttpResponseRedirect(reverse("addIneventoryManual"))

    elif request.method =="GET":
        getExcelTemplate(request)
        return render(request,"addExcelStock.html",{})

def getExcelTemplate(request):
    dest_filename = "dashboard/static/customerTemplate.xlsx"

    with open(dest_filename,"w") as f:
        f.close()

    wb = Workbook()

    ws = wb.active

    ws['A1'] = "Item (Do not touch number between [])"
    ws['B1'] = "Count"

    items = Item.objects.all()
    items = sorted(items,key=lambda x:x.name)
    rowNum = 2
    for item in items:
        ws["A" + str(rowNum)] = item.name + "[" + str(item.pk) + "]"
    wb.save(filename=dest_filename)

def userLogin(request):
    if request.method=="GET":
        return render(request,"login.html",{})

    elif request.method=="POST":
        username = request.POST["username"]
        password = request.POST["password"]

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request,user)
            return HttpResponseRedirect(reverse("inventoryView"))
        else:
            return HttpResponseRedirect(reverse("login"))



def register(request):
    if request.method=="GET":
        return render(request,"register.html",{})
    else:
        firstName = request.POST["firstName"]
        lastName = request.POST["lastName"]
        company = request.POST["companyName"]
        address = request.POST["address"]
        password = request.POST["password"]
        email = request.POST["email"]
        username = request.POST["username"]

        user = User()
        user.username = username
        user.first_name = firstName
        user.last_name = lastName
        user.email = email
        user.set_password(password)
        user.save()

        login(request,user)
        return HttpResponseRedirect(reverse("inventoryView"))

def userLogout(request):
    logout(request)
    return HttpResponseRedirect(reverse("dashboard"))

def generateReport(request):
    invoices = Invoice.objects.all()
    earned = 0
    cost = 0
    quantities = defaultdict(int)
    for invoice in invoices:
        for invoiceItem in invoice.invoiceitem_set.all():
            cost += invoiceItem.quantity * invoiceItem.item.costPrice
            quantities[invoiceItem.item.pk] += invoiceItem.quantity
            print(quantities)
        earned += invoice.finalSale
    profit = earned - cost

    maxQuantity = -1
    maxQuantityPK = None
    for item in quantities:
        if quantities[item] > maxQuantity:
            maxQuantity = quantities[item]
            maxQuantityPK = item
    print(quantities)
    mostPopularItem = Item.objects.get(pk=maxQuantityPK)

    #mostboughtgood
    quantitiesBought = defaultdict(int)
    for item in Item.objects.all():
        for stockupdate in item.stockupdate_set.all():
            quantitiesBought[item.pk] += stockupdate.count
            print(quantitiesBought)
    maxBought = -1
    maxBoughtPK = None
    for pk in quantitiesBought:
        if quantitiesBought[pk] > maxBought:
            maxBought = quantitiesBought[pk]
            maxBoughtPK = pk

    mostBoughtGood = Item.objects.get(pk=maxBoughtPK)

    boughtProfit = maxBought * (mostBoughtGood.salesPrice - mostBoughtGood.costPrice)
    salesProfit = maxQuantity * (mostPopularItem.salesPrice - mostPopularItem.salesPrice)

    

    return render(request, "monetary.html", {"mostBoughtItem":mostBoughtGood,"mostBoughtQuantity":maxBought,
                                             "mostSoldItem":mostPopularItem,"mostSoldQuantity":maxQuantity,
                                             "profit":profit,"boughtProfit":boughtProfit,"soldPRofit":salesProfit})


